#!/usr/bin/env python3
"""Compare LLM selector predictions against actual CI failures (Lane 3).

Usage:
    python compare_selector_vs_ci.py <pr_number> --ci-evidence <path> --selector-replay <path>
"""

import argparse
import json
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Set, List, Tuple
import re

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    OPENPYXL_AVAILABLE = True
    # Colors (only used when openpyxl is available)
    GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    HEADER_FONT = Font(color="FFFFFF", bold=True)
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import yaml
    YAML_AVAILABLE = True
except ImportError:
    YAML_AVAILABLE = False


def _removeprefix(s: str, prefix: str) -> str:
    """Backport of str.removeprefix for Python < 3.9."""
    if s.startswith(prefix):
        return s[len(prefix):]
    return s


def load_json(path: str) -> Dict:
    """Load JSON from a file."""
    try:
        return json.loads(Path(path).read_text())
    except Exception as e:
        print(f"Error loading {path}: {e}", file=sys.stderr)
        return {}


def build_test_to_job_mapping_from_yaml() -> Dict[str, Set[str]]:
    """Build a mapping from test file paths to job names by parsing .buildkite/test_areas/*.yaml files.

    Returns a dict mapping test_path → set of job names that run that test.
    Uses simple regex-based parsing to avoid yaml dependency.
    """
    mapping: Dict[str, Set[str]] = {}

    test_areas_dir = Path(".buildkite/test_areas")
    if not test_areas_dir.exists():
        return mapping

    for yaml_file in test_areas_dir.glob("*.yaml"):
        try:
            content = yaml_file.read_text()

            # Extract job labels and their pytest commands using simple regex
            # Pattern: "- label: <name>" followed by "pytest <path>"
            current_label = None
            for line in content.splitlines():
                # Check for job label
                label_match = re.match(r'^\s*-\s*label:\s*(.+)$', line)
                if label_match:
                    current_label = label_match.group(1).strip()
                    continue

                # Check for pytest commands
                if current_label and 'pytest' in line:
                    # Extract pytest test path
                    # Examples: "pytest -v -s quantization/"
                    #           "pytest model_executor -m"
                    tokens = line.split()
                    for i, token in enumerate(tokens):
                        if token == 'pytest':
                            # Look for the test path in the next few tokens
                            for j in range(i + 1, min(i + 10, len(tokens))):
                                if j >= len(tokens):
                                    break
                                arg = tokens[j]
                                # Skip flags and options
                                if arg.startswith('-'):
                                    continue
                                # Skip known non-path arguments
                                if arg in ('--ignore', '-m', '--ignore-glob', 'tests'):
                                    continue
                                # If it looks like a path
                                if '/' in arg or arg.endswith('.py') or (
                                    re.match(r'^[\w_]+$', arg) and not arg.isupper()
                                ):
                                    test_path = arg.rstrip('/')
                                    # Normalize path (remove tests/ prefix if present)
                                    test_path_norm = _removeprefix(test_path, "tests/")
                                    mapping.setdefault(test_path_norm, set()).add(current_label)

                                    # Also add parent directories for directory-level matches
                                    parts = test_path_norm.split("/")
                                    for k in range(1, len(parts) + 1):
                                        parent = "/".join(parts[:k])
                                        mapping.setdefault(parent, set()).add(current_label)
                                    break
                            break

        except Exception as e:
            print(f"Warning: Failed to parse {yaml_file}: {e}", file=sys.stderr)
            continue

    return mapping


def classify_selections(selected_tests: List[Dict], failed_tests: List[Dict]) -> Tuple[List, List, List]:
    """Classify into true positives, false negatives, false positives.

    A selected test matches a failure if:
    - Exact match: same test file and function
    - File match: selected test file contains the failed test function
    - Directory match: selected directory contains the failed test file
    """
    true_positives, false_negatives, false_positives = [], [], []
    matched = set()

    def test_matches(sel_id: str, fail_id: str) -> bool:
        """Check if selected test matches a failure."""
        sel = sel_id.rstrip("/")
        fail = fail_id.rstrip("/")

        # Exact match
        if sel == fail:
            return True

        # Selected is directory containing failed test
        if sel_id.endswith("/"):
            if fail.startswith(sel) or fail.startswith(sel.rstrip("/")):
                return True

        # Selected file contains failed function (e.g., test_file.py matches test_file.py::test_fn)
        if "::" in fail_id:
            fail_file = fail_id.split("::")[0]
            if sel == fail_file or sel.endswith("/" + fail_file.split("/")[-1]):
                return True

        # Directory containment (selected dir is parent of failed file)
        sel_parts = sel.replace("\\", "/").split("/")
        fail_parts = fail.replace("\\", "/").split("/")
        if len(sel_parts) < len(fail_parts):
            if fail_parts[:len(sel_parts)] == sel_parts:
                return True

        return False

    for selected in selected_tests:
        sel_id = selected.get("identifier", "")
        found = False
        for i, failed in enumerate(failed_tests):
            fail_id = failed.get("identifier", "")
            if test_matches(sel_id, fail_id):
                found = True
                matched.add(i)
                true_positives.append({"selected_test": sel_id, "matched_failure": fail_id, "reason": selected.get("reason", "")})
                break
        if not found:
            false_positives.append({"selected_test": sel_id, "reason": selected.get("reason", "")})

    for i, failed in enumerate(failed_tests):
        if i not in matched:
            false_negatives.append({
                "failed_test": failed.get("identifier", ""),
                "job_name": failed.get("job_name", ""),
                "why_missed": "not_in_candidate_mapping"
            })

    return true_positives, false_negatives, false_positives


def compute_metrics(tp: List, fn: List, fp: List) -> Dict:
    """Compute evaluation metrics."""
    coverage = len(tp) / (len(tp) + len(fn)) if (tp or fn) else 1.0
    precision = len(tp) / (len(tp) + len(fp)) if (tp or fp) else 1.0
    return {
        "true_positives": len(tp), "false_negatives": len(fn), "false_positives": len(fp),
        "coverage_rate": round(coverage, 3), "precision_rate": round(precision, 3),
    }



def generate_excel(pr_number: str, ci_evidence: Dict, selector_replay: Dict, buildkite_jobs: List,
                   metrics: Dict, tp: List, fn: List, fp: List) -> bytes:
    """Generate Excel file with 2 sheets."""
    if not OPENPYXL_AVAILABLE:
        return None

    wb = Workbook()

    # Sheet 1: Lane Comparison
    ws1 = wb.active
    ws1.title = "Lane Comparison"
    ws1.append(["Category", "Test/Job", "Source", "Selected?", "Failed?", "Reason"])
    for c in ws1[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    row = 2
    for t in tp:
        ws1.append(["TRUE POSITIVE", t['selected_test'], "Both", "YES", "YES", t.get('reason', '')])
        for c in ws1[row]:
            c.fill = GREEN_FILL
        row += 1

    for f in fn:
        ws1.append(["FALSE NEGATIVE", f['failed_test'], "CI Only", "NO", "YES", f.get('why_missed', '')])
        for c in ws1[row]:
            c.fill = RED_FILL
        row += 1

    for f in fp:
        ws1.append(["FALSE POSITIVE", f['selected_test'], "LLM Only", "YES", "NO", f.get('reason', '')])
        for c in ws1[row]:
            c.fill = YELLOW_FILL
        row += 1

    # Sheet 2: Buildkite CI Tests
    ws2 = wb.create_sheet(title="Buildkite CI Tests")
    ws2.append(["Job Name", "State", "Failed?", "Exit Status"])
    for c in ws2[1]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT

    for job in buildkite_jobs:
        if job.get('state') == 'blocked':
            continue
        state = job.get('state', 'unknown').upper()
        failed = "YES" if state == "FAILED" else "NO"
        ws2.append([job.get('name', ''), state, failed, job.get('exit_status', 'N/A')])
        if state == "FAILED":
            for c in ws2[ws2.max_row]:
                c.fill = RED_FILL
        else:
            for c in ws2[ws2.max_row]:
                c.fill = GREEN_FILL

    from io import BytesIO
    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out.read()


def derive_failed_tests_from_jobs(jobs_failed: List[Dict], yaml_mapping: Dict[str, Set[str]]) -> List[Dict]:
    """When individual test names aren't available, derive failed test paths from failed job names.

    Reverses the YAML test-to-job mapping to find which test paths each failed job runs.
    Returns a list of failed test dicts compatible with failed_test_list format.
    """
    # Build reverse mapping: job_name -> set of test_paths
    job_to_tests: Dict[str, Set[str]] = {}
    for test_path, jobs in yaml_mapping.items():
        for job in jobs:
            job_to_tests.setdefault(job, set()).add(test_path)

    failed_job_names = {j.get("name", "") for j in jobs_failed}
    derived: List[Dict] = []
    seen: Set[str] = set()

    for job_name in failed_job_names:
        test_paths = job_to_tests.get(job_name, set())
        if test_paths:
            for test_path in test_paths:
                full_path = f"tests/{test_path}" if not test_path.startswith("tests/") else test_path
                if full_path not in seen:
                    seen.add(full_path)
                    derived.append({
                        "identifier": full_path,
                        "file": full_path,
                        "job_name": job_name,
                        "state": "failed",
                        "source": "job_level_fallback",
                    })
        else:
            # Job has no YAML mapping — add a job-level entry so it shows up in the table
            label = f"[job] {job_name}"
            if label not in seen:
                seen.add(label)
                derived.append({
                    "identifier": label,
                    "file": label,
                    "job_name": job_name,
                    "state": "failed",
                    "source": "job_level_fallback_unmapped",
                })

    return derived


def main():
    parser = argparse.ArgumentParser(description="Compare LLM selector vs CI failures")
    parser.add_argument("pr_number", type=str)
    parser.add_argument("--ci-evidence", required=True)
    parser.add_argument("--selector-replay", required=True)
    parser.add_argument("--output-dir", default=None)
    args = parser.parse_args()

    pr_number = args.pr_number
    now = datetime.now(timezone.utc)

    print(f"Comparing selector vs CI for PR #{pr_number}...", file=sys.stderr)

    # Load inputs
    ci_evidence = load_json(args.ci_evidence)
    selector_replay = load_json(args.selector_replay)

    if not ci_evidence or not selector_replay:
        print("Error: Could not load inputs", file=sys.stderr)
        sys.exit(1)

    # Extract data — prefer individual test names; fall back to job-level when logs unavailable
    failed_tests = ci_evidence.get("failed_test_list", [])
    selected_tests = selector_replay.get("llm_selected_tests", [])

    yaml_mapping = build_test_to_job_mapping_from_yaml()

    if not failed_tests and ci_evidence.get("jobs_failed"):
        derived = derive_failed_tests_from_jobs(ci_evidence["jobs_failed"], yaml_mapping)
        if derived:
            print(f"  Note: no individual test names in evidence; derived {len(derived)} test paths "
                  f"from {len(ci_evidence['jobs_failed'])} failed jobs", file=sys.stderr)
            failed_tests = derived

    # Classify
    tp, fn, fp = classify_selections(selected_tests, failed_tests)
    metrics = compute_metrics(tp, fn, fp)

    # Output dir
    output_dir = Path(args.output_dir) if args.output_dir else Path(f".buildkite/test_selection_skills/evaluation/pr_{pr_number}")
    output_dir.mkdir(parents=True, exist_ok=True)

    # Write evaluation report
    report = {
        "pr_number": pr_number,
        "evaluation_status": "success",
        "summary": metrics,
        "true_positive_details": tp,
        "false_negative_details": fn,
        "false_positive_details": fp,
        "notes": [f"Generated on {now.isoformat()}"],
    }
    (output_dir / "evaluation_report.json").write_text(json.dumps(report, indent=2))

    # Write report.md — single output file (follows report.template.md structure)
    builds = ci_evidence.get('buildkite_builds', [])
    # Prefer the main vllm/ci pipeline build over subsidiary pipelines (e.g. intel-ci)
    main_build = next(
        (b for b in builds if b.get('pipeline', '').rstrip('/') == '/ci'),
        builds[0] if builds else {}
    )
    build_number = main_build.get('build_number', 'unknown')
    build_url = f"https://buildkite.com/vllm/ci/builds/{build_number}" if build_number != 'unknown' else ""

    # Fall back to GitHub-sourced build links when Buildkite API was unavailable
    if build_number == 'unknown':
        status_jobs = ci_evidence.get('buildkite_jobs_from_statuses', [])
        main_status = next(
            (j for j in status_jobs if j.get('pipeline', '') == 'ci'),
            status_jobs[0] if status_jobs else {}
        )
        if main_status:
            build_number = main_status.get('build_number', 'unknown')
            build_url = main_status.get('build_url', '')

    # Last resort: link to the GitHub PR itself
    if build_number == 'unknown':
        repo = ci_evidence.get('repo', 'vllm-project/vllm')
        build_number = f"PR #{pr_number}"
        build_url = f"https://github.com/{repo}/pull/{pr_number}"

    # Group failures by job for use in multiple sections
    tests_by_job: Dict = {}
    for t in failed_tests:
        tests_by_job.setdefault(t.get('job_name', 'Unknown Job'), []).append(
            t.get('identifier', t.get('test_name', '')))

    summary_stats = ci_evidence.get('summary_stats', {})
    jobs_run = ci_evidence.get('jobs_run', [])
    n_passed_jobs = sum(1 for j in jobs_run if j.get('state') == 'passed')
    n_failed_jobs = sum(1 for j in jobs_run if j.get('state') == 'failed')
    n_blocked_jobs = sum(1 for j in jobs_run if j.get('state') == 'blocked')

    report_md = f"# PR #{pr_number} — Test Selection Evaluation\n\n"
    report_md += f"> **Build**: [{build_number}]({build_url}) · **Date**: {now.strftime('%Y-%m-%d')}\n\n"

    # --- Metrics ---
    report_md += "## Metrics\n\n"
    report_md += "| Metric | Value |\n|--------|-------|\n"
    report_md += f"| **Recall** | **{metrics['coverage_rate']:.1%}** |\n"
    report_md += f"| **Precision** | **{metrics['precision_rate']:.1%}** |\n"
    report_md += f"| True Positives | {len(tp)} |\n"
    report_md += f"| False Negatives (CI failed, LLM missed) | {len(fn)} |\n"
    report_md += f"| False Positives (LLM selected, passed) | {len(fp)} |\n\n"

    # --- Full CI picture ---
    report_md += "## CI Test Results\n\n"

    if summary_stats:
        total = summary_stats.get('failed', 0) + summary_stats.get('passed', 0) + summary_stats.get('skipped', 0)
        report_md += (f"| Failed | Passed | Skipped | Total |\n"
                      f"|--------|--------|---------|-------|\n"
                      f"| {summary_stats.get('failed', '?')} "
                      f"| {summary_stats.get('passed', '?')} "
                      f"| {summary_stats.get('skipped', '?')} "
                      f"| {total} |\n\n")

    if jobs_run:
        report_md += (f"**CI jobs:** {n_passed_jobs} passed · "
                      f"{n_failed_jobs} failed · {n_blocked_jobs} blocked")

    # --- Build test-to-job mapping first (used by multiple sections) ---
    # Build a comprehensive mapping from test file paths to job names
    # First, try to load from YAML definitions (static mapping)
    yaml_test_to_jobs = yaml_mapping  # already built above

    # Also extract from CI evidence (runtime mapping from actual failures)
    ci_test_to_jobs: Dict[str, Set[str]] = {}
    all_ci_tests = ci_evidence.get("failed_test_list", [])

    for test_item in all_ci_tests:
        test_path = test_item.get("identifier", "")
        job_name = test_item.get("job_name", "")
        if test_path and job_name:
            # Extract just the file path (remove ::test_function if present)
            test_file = test_path.split("::")[0] if "::" in test_path else test_path
            # Strip "tests/" prefix if present for matching
            test_file_normalized = _removeprefix(test_file, "tests/")
            ci_test_to_jobs.setdefault(test_file_normalized, set()).add(job_name)

    # Merge both mappings (CI evidence takes precedence as it's runtime data)
    combined_mapping = yaml_test_to_jobs.copy()
    for test_path, jobs in ci_test_to_jobs.items():
        combined_mapping.setdefault(test_path, set()).update(jobs)

    # --- Comparison Table: Row-by-row format ---
    # Lists ALL test files from CI (YAML mapping) with LLM Selected and CI Result columns

    # Build LLM selected set (normalized)
    llm_selected_tests = {t.get('identifier', '').rstrip('/') for t in selected_tests if t.get('identifier')}

    # Build CI test set - ALL tests from YAML mapping (not just relevant ones)
    # This ensures we show all tests that run in CI
    ci_all_tests = {f"tests/{test_path}".rstrip('/') for test_path in combined_mapping.keys()}

    # Build CI failure set
    ci_failed_tests = {f.get('failed_test', '').rstrip('/') for f in fn if f.get('failed_test')}
    ci_all_tests.update(ci_failed_tests)

    # Include ALL CI tests plus any LLM-selected tests that might not be in the mapping
    all_test_files = ci_all_tests | llm_selected_tests

    # Add test file count to the CI jobs line
    if jobs_run:
        report_md += f" · **{len(all_test_files)} test files**\n\n"

    report_md += "## LLM vs CI Comparison Table\n\n"

    # Build rows for row-by-row table (matching PR #37505 format)
    rows = []
    for test_file in sorted(all_test_files):
        llm_mark = "✓" if test_file in llm_selected_tests else "✗"

        if test_file in ci_failed_tests:
            ci_mark = "✗ Failed"
        else:
            ci_mark = "✓ Passed"

        rows.append((test_file, llm_mark, ci_mark))

    if rows:
        # Calculate column widths for fixed-width alignment (matching PR #37505)
        max_test_width = max([len(r[0]) for r in rows] + [len("Test File")])
        llm_width = len("LLM Selected")
        ci_width = len("CI Result")

        report_md += "```\n"

        # Header (matching PR #37505 format)
        header = f"{'Test File'.ljust(max_test_width)}  {'LLM Selected'.ljust(llm_width)}  {'CI Result'.ljust(ci_width)}"
        report_md += header + "\n"
        report_md += "=" * len(header) + "\n"

        # Data rows
        for test_file, llm_mark, ci_mark in rows:
            report_md += f"{test_file.ljust(max_test_width)}  {llm_mark.ljust(llm_width)}  {ci_mark.ljust(ci_width)}\n"

        report_md += "```\n\n"

        # Summary counts
        llm_count = sum(1 for _, llm, _ in rows if llm == "✓")
        ci_failed_count = sum(1 for _, _, ci in rows if "Failed" in ci)
        report_md += f"**Summary:** {len(rows)} total tests, {llm_count} LLM selected, {ci_failed_count} CI failed\n\n"

    # Per-job breakdown: failed jobs with named tests
    if tests_by_job:
        report_md += "### Failed jobs\n\n"
        for job_name, job_tests in sorted(tests_by_job.items()):
            report_md += f"**❌ {job_name}** — {len(job_tests)} test(s) failed\n\n"
            for test in sorted(job_tests):
                report_md += f"- `{test}`\n"
            report_md += "\n"

    # Passing jobs with test file lists (matching PR #37505 format)
    passing_jobs = [j for j in jobs_run
                    if j.get('state') == 'passed' and j.get('name', '').strip()]
    if passing_jobs:
        report_md += "### Passing jobs\n\n"

        # Build reverse mapping: job_name -> test_paths
        job_to_tests: Dict[str, Set[str]] = {}
        for test_path, jobs in combined_mapping.items():
            for job in jobs:
                job_to_tests.setdefault(job, set()).add(test_path)

        for j in sorted(passing_jobs, key=lambda x: x.get('name', '')):
            job_name = j['name']
            report_md += f"**✅ {job_name}**\n\n"

            # Show test files this job ran
            test_files = job_to_tests.get(job_name, set())
            if test_files:
                # Add tests/ prefix back for display
                for test_file in sorted(test_files):
                    report_md += f"- `tests/{test_file}`\n"
            else:
                report_md += "- (test paths unknown)\n"
            report_md += "\n"

    # --- LLM selections with job mapping (matching PR #37505 format) ---
    report_md += f"## LLM Selections — {len(selected_tests)} target(s)\n\n"

    for t in selected_tests:
        caught = any(tp_item['selected_test'] == t.get('identifier') for tp_item in tp)
        status = "✅" if caught else "➖"
        test_id = t.get('identifier', '')

        # Try to find matching jobs for this test selection
        # Normalize the test identifier for matching
        test_normalized = _removeprefix(test_id, "tests/").rstrip("/")

        # Find jobs that would run this test
        matching_jobs = set()

        # Direct match
        if test_normalized in combined_mapping:
            matching_jobs.update(combined_mapping[test_normalized])

        # Check for directory matches (selected dir contains tests in mapping)
        is_directory = test_id.endswith("/")

        if is_directory:
            # Directory selection - need to check if this directory would be run by any parent or exact match
            # Example: test_normalized = "v1/simple_kv_offload"
            #          should match "v1" or "v1/kv_offload" or exact "v1/simple_kv_offload" in mapping

            # Check exact match first
            if test_normalized in combined_mapping:
                matching_jobs.update(combined_mapping[test_normalized])

            # Check parent directories that would run this directory
            parts = test_normalized.split("/")
            for i in range(len(parts)):
                prefix = "/".join(parts[:i+1])
                if prefix in combined_mapping:
                    matching_jobs.update(combined_mapping[prefix])

            # Also check if any mapping entries are subdirectories that match
            for test_path, jobs in combined_mapping.items():
                if test_path.startswith(test_normalized + "/") or test_path == test_normalized:
                    matching_jobs.update(jobs)
        else:
            # File selection - need to check if this file would be run by any directory-level job
            # Example: test_normalized = "quantization/test_fp8.py"
            #          should match "quantization" or "quantization/" in mapping

            # Split into parts to check parent directories
            parts = test_normalized.split("/")
            for i in range(len(parts)):
                prefix = "/".join(parts[:i+1])
                if prefix in combined_mapping:
                    matching_jobs.update(combined_mapping[prefix])

        # Format output with test file and jobs listed underneath
        report_md += f"### {status} `{test_id}`\n\n"
        report_md += f"**Reason:** {t.get('reason', 'N/A')}\n\n"

        if matching_jobs:
            report_md += f"**Jobs ({len(matching_jobs)}):**\n\n"
            for job in sorted(matching_jobs):
                report_md += f"- 🔧 {job}\n"
        else:
            report_md += "**Jobs:** (unknown)\n"
        report_md += "\n"

    # --- Gap analysis (auto-derived from test paths) ---
    report_md += "## Gap Analysis\n\n"
    if fn:
        # Extract top-level test directory from a test path
        def top_dir(path: str) -> str:
            p = path.replace("\\", "/").split("::")[0]  # strip ::test_fn
            p = _removeprefix(p, "tests/")
            return p.split("/")[0]

        failed_dirs = {top_dir(f.get('failed_test', '')) for f in fn if f.get('failed_test')}
        selected_dirs = {top_dir(t.get('identifier', '')) for t in selected_tests if t.get('identifier')}

        missed_dirs = failed_dirs - selected_dirs

        report_md += "**Why the LLM missed:**\n"
        if missed_dirs:
            report_md += (f"- LLM selections covered `{'`, `'.join(sorted(selected_dirs))}` "
                          f"but failures occurred in `{'`, `'.join(sorted(missed_dirs))}`\n")
        for f in fn:
            test = f.get('failed_test', '')
            job = f.get('job_name', '')
            report_md += f"- `{test}` (job: {job}) was not covered by any selection\n"

        report_md += "\n**To improve coverage:**\n"
        for d in sorted(missed_dirs):
            report_md += f"- Add `tests/{d}/` (or relevant sub-paths) to selections when related code changes\n"
        report_md += "\n"
    else:
        report_md += "LLM caught all CI failures — no gap to analyse.\n\n"

    report_md += f"---\n*Generated: {now.strftime('%Y-%m-%d %H:%M UTC')}*\n"

    (output_dir / "report.md").write_text(report_md)

    print(f"Coverage: {metrics['coverage_rate']:.1%}, Precision: {metrics['precision_rate']:.1%}", file=sys.stderr)
    print(f"TP: {len(tp)}, FN: {len(fn)}, FP: {len(fp)}", file=sys.stderr)
    print(f"Generated comprehensive reports in: {output_dir}", file=sys.stderr)

    return 0


if __name__ == "__main__":
    sys.exit(main())
